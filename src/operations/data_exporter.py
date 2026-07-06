"""Data export operations (CSV, JSON, Excel)."""

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging

from ..core.exceptions import ExportError
from ..core.soql_validator import SOQLSanitizer

logger = logging.getLogger(__name__)

# Check for optional dependencies
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class DataExporter:
    """Exports Salesforce data to various formats."""

    def __init__(self, output_dir: Path):
        """Initialize exporter.

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Initialized DataExporter with output dir: {output_dir}")

    def export_csv(
        self,
        headers: List[str],
        rows: List[Dict[str, str]],
        filename: str,
    ) -> Path:
        """Export data to CSV file.

        Args:
            headers: Column headers
            rows: List of row dictionaries
            filename: Output filename (without .csv)

        Returns:
            Path to created file

        Raises:
            ExportError: If export fails
        """
        try:
            filepath = self.output_dir / f"{filename}.csv"
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)
            logger.info(f"Exported {len(rows)} rows to {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            raise ExportError("CSV", str(self.output_dir), str(e))

    def export_json(
        self,
        records: List[Dict[str, Any]],
        filename: str,
        flat: bool = False,
        indent: int = 2,
    ) -> Path:
        """Export data to JSON file.

        Args:
            records: List of record dictionaries
            filename: Output filename (without .json)
            flat: If True, use flat row format; else use native structure
            indent: JSON indentation level

        Returns:
            Path to created file

        Raises:
            ExportError: If export fails
        """
        try:
            filepath = self.output_dir / f"{filename}.json"

            # Clean metadata if exporting native structure
            if not flat:
                records = [
                    SOQLSanitizer.clean_metadata(r) for r in records
                ]

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=indent)

            logger.info(f"Exported {len(records)} records to {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            raise ExportError("JSON", str(self.output_dir), str(e))

    def export_xlsx(
        self,
        headers: List[str],
        rows: List[Dict[str, str]],
        filename: str,
    ) -> Path:
        """Export data to Excel (.xlsx) file.

        Args:
            headers: Column headers
            rows: List of row dictionaries
            filename: Output filename (without .xlsx)

        Returns:
            Path to created file

        Raises:
            ExportError: If export fails or openpyxl not installed
        """
        if not XLSX_AVAILABLE:
            raise ExportError(
                "XLSX",
                str(self.output_dir),
                "openpyxl not installed. Install with: pip install openpyxl",
            )

        try:
            filepath = self.output_dir / f"{filename}.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
                    max(20, len(header))
                )

            wb.save(filepath)
            logger.info(f"Exported {len(rows)} rows to {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"XLSX export failed: {e}")
            raise ExportError("XLSX", str(self.output_dir), str(e))
